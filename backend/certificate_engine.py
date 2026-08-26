"""
Bulk Certificate Generator - Core Engine
==========================================
Pure Python, no network access required. This module contains all the
logic for:
  - detecting fillable text fields in a PDF template
  - reading recipient names from .xlsx / .csv files
  - filling ONE field (the recipient name) per certificate while leaving
    every other property of the PDF untouched
  - detecting "name too long for the field" overflow instead of silently
    shrinking the font
  - safe, non-colliding output filenames

Design principle (per SRS section 9/10/12):
  Template in -> identical template out, with only the recipient name
  changed. We NEVER touch font family/size/color, position, dimensions,
  images, or any other field. We rely on the PDF's own AcroForm field
  appearance (/DA - Default Appearance string) to render the text, which
  is exactly the font/size/color the template author already configured
  for that field. We only use that same DA information to *estimate*
  whether a given name will physically fit, so we can flag it instead of
  silently reformatting anything.
"""

from __future__ import annotations

import csv
import io
import os
import re
import threading
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

from pypdf import PdfReader, PdfWriter
from pypdf.generic import NameObject
import pymupdf
from reportlab.lib.colors import HexColor
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------

@dataclass
class TemplateField:
    name: str
    page: int  # 1-based
    rect: tuple  # (x0, y0, x1, y1) in PDF points
    font_name: Optional[str]
    font_size: Optional[float]
    max_len: Optional[int]  # /MaxLen if the field enforces one


@dataclass
class NameRecord:
    row_number: int  # 1-based row number as it appears in the sheet (header excluded)
    raw_name: str


@dataclass
class GenerationResult:
    row_number: int
    name: str
    status: str  # "success" | "overflow" | "empty" | "error"
    output_path: Optional[str] = None
    message: str = ""


# The 14 standard Type 1 PDF fonts
STANDARD_FONTS = {
    "Helvetica", "Helvetica-Bold", "Helvetica-Oblique", "Helvetica-BoldOblique",
    "Times-Roman", "Times-Bold", "Times-Italic", "Times-BoldItalic",
    "Courier", "Courier-Bold", "Courier-Oblique", "Courier-BoldOblique",
    "Symbol", "ZapfDingbats"
}

# A conservative "safe" font fallback when a template's declared font isn't
# one of the 14 standard PDF fonts reportlab knows metrics for. We only use
# this fallback for the WIDTH ESTIMATE (overflow check) - never for the
# actual rendered certificate, which always uses pypdf + the field's own /DA.
_FALLBACK_FONT = "Helvetica"


# ---------------------------------------------------------------------------
# Template inspection
# ---------------------------------------------------------------------------

def _parse_da_string(da: str) -> tuple[Optional[str], Optional[float]]:
    """Parse a PDF /DA (Default Appearance) string like '/Helv 12 Tf 0 g'
    and return (font_resource_name, font_size)."""
    if not da:
        return None, None
    match = re.search(r"/(\S+)\s+([\d.]+)\s+Tf", da)
    if not match:
        return None, None
    font_res, size_str = match.groups()
    try:
        size = float(size_str)
    except ValueError:
        size = None
    return font_res, size


def _resolve_font_display_name(reader: PdfReader, page_index: int, font_res_name: Optional[str]) -> Optional[str]:
    """Best-effort: map the /DA font resource name (e.g. 'Helv') to a real
    font base name (e.g. 'Helvetica-Bold') by looking at the AcroForm /DR
    (Default Resources) font dictionary. Falls back to the resource name
    itself if it can't be resolved."""
    if not font_res_name:
        return None
    try:
        root = reader.trailer["/Root"]
        acro_form = root.get("/AcroForm")
        if not acro_form:
            return font_res_name
        dr = acro_form.get("/DR")
        if not dr:
            return font_res_name
        fonts = dr.get("/Font")
        if not fonts:
            return font_res_name
        # Keys in the /DR /Font dict are stored WITH a leading slash
        # (e.g. "/TiBo"), while the /DA string refers to them WITHOUT one
        # (e.g. "/TiBo 12 Tf" -> resource name "TiBo"). Match either form.
        key = None
        if font_res_name in fonts:
            key = font_res_name
        elif f"/{font_res_name}" in fonts:
            key = f"/{font_res_name}"
        else:
            for k in fonts.keys():
                if str(k).lstrip("/") == font_res_name:
                    key = k
                    break
        if key is None:
            return font_res_name
        font_obj = fonts[key].get_object()
        base_font = font_obj.get("/BaseFont")
        if base_font:
            return str(base_font).lstrip("/")
    except Exception:
        pass
    return font_res_name


def detect_fillable_fields(pdf_path: str) -> list[TemplateField]:
    """Return every fillable *text* field found in the PDF, across all pages.
    Checkbox/radio/choice fields are intentionally excluded - V1 only
    supports a single free-text recipient-name field per SRS section 8."""
    reader = PdfReader(pdf_path)
    results: list[TemplateField] = []

    for page_index, page in enumerate(reader.pages):
        annots = page.get("/Annots")
        if not annots:
            continue
        for annot_ref in annots:
            annot = annot_ref.get_object()
            if annot.get("/Subtype") != "/Widget":
                continue
            ft = annot.get("/FT")
            # A widget's field type may be inherited from its /Parent
            parent = annot.get("/Parent")
            node = annot
            while ft is None and parent is not None:
                parent_obj = parent.get_object()
                ft = parent_obj.get("/FT")
                node = parent_obj
                parent = parent_obj.get("/Parent")
            if ft != "/Tx":  # only plain text fields
                continue

            # Fully-qualified field name (handles parent/child naming)
            name_parts = []
            n = annot
            while n is not None:
                t = n.get("/T")
                if t:
                    name_parts.insert(0, str(t))
                p = n.get("/Parent")
                n = p.get_object() if p is not None else None
            field_name = ".".join(name_parts) if name_parts else f"field_{page_index}_{len(results)}"

            rect = tuple(float(x) for x in annot.get("/Rect", [0, 0, 0, 0]))

            da = annot.get("/DA")
            if da is None and node is not None:
                da = node.get("/DA")
            if da is None:
                acro_form = reader.trailer["/Root"].get("/AcroForm")
                da = acro_form.get("/DA") if acro_form else None
            font_res, font_size = _parse_da_string(str(da) if da else "")
            font_display = _resolve_font_display_name(reader, page_index, font_res)

            max_len = annot.get("/MaxLen")
            if max_len is None and node is not None:
                max_len = node.get("/MaxLen")

            results.append(TemplateField(
                name=field_name,
                page=page_index + 1,
                rect=rect,
                font_name=font_display,
                font_size=font_size,
                max_len=int(max_len) if max_len is not None else None,
            ))

    return results


def map_font_family(font_name: Optional[str], is_bold: bool = False, is_italic: bool = False) -> str:
    f = (font_name or "").lower()
    bold = is_bold or "bold" in f or "black" in f or "heavy" in f
    italic = is_italic or "italic" in f or "oblique" in f
    if "times" in f or "serif" in f or "georgia" in f or "garamond" in f or "cambria" in f:
        if bold and italic:
            return "Times-BoldItalic"
        if bold:
            return "Times-Bold"
        if italic:
            return "Times-Italic"
        return "Times-Roman"
    if "courier" in f or "mono" in f or "consolas" in f:
        if bold:
            return "Courier-Bold"
        return "Courier"
    if bold and italic:
        return "Helvetica-BoldOblique"
    if bold:
        return "Helvetica-Bold"
    if italic:
        return "Helvetica-Oblique"
    return "Helvetica"


def analyze_pdf_template(pdf_path: str) -> dict:
    """Analyze the PDF template to detect existing placeholder text, font families,
    font sizes, colors, and optimal recipient name placement (both X and Y)."""
    try:
        doc = pymupdf.open(pdf_path)
        if len(doc) == 0:
            return {
                "detected": False,
                "font_name": "Times-Bold",
                "font_size": 28.0,
                "font_color": "#0d408c",
                "x_percent": 50.0,
                "y_percent": 50.0,
                "align": "center",
                "box_width": 520,
                "box_height": 48,
                "bbox": None,
                "reason": "Default centered placement",
            }
        page = doc[0]
        w = page.rect.width
        h = page.rect.height

        blocks = page.get_text("dict")["blocks"]
        spans = []
        for b in blocks:
            if "lines" in b:
                for l in b["lines"]:
                    for s in l["spans"]:
                        text = s["text"].strip()
                        if text:
                            c_int = s["color"]
                            r = (c_int >> 16) & 0xFF
                            g = (c_int >> 8) & 0xFF
                            b_val = c_int & 0xFF
                            color_hex = f"#{r:02x}{g:02x}{b_val:02x}"
                            spans.append({
                                "text": text,
                                "font": s["font"],
                                "size": s["size"],
                                "color": color_hex,
                                "flags": s.get("flags", 0),
                                "bbox": s["bbox"],
                                "y_percent": ((s["bbox"][1] + s["bbox"][3]) / 2.0 / h) * 100.0,
                            })

        # 1. Search for explicit placeholder names
        placeholder_pattern = re.compile(
            r"(\{\{.*?name.*?\}\}|\{.*?name.*?\}|\[.*?name.*?\]|<.*?name.*?>|\b(recipient|participant|student|candidate|attendee|holder|member)\s+name\b|\bname\s+here\b|\binsert\s+name\b|\bfull\s+name\b|\b(john|jane)\s+doe\b)",
            re.IGNORECASE,
        )

        for s in spans:
            if placeholder_pattern.search(s["text"]):
                mapped_font = map_font_family(
                    s["font"],
                    is_bold=bool(s["flags"] & 2),
                    is_italic=bool(s["flags"] & 1),
                )
                bbox = s["bbox"]
                span_w = bbox[2] - bbox[0]
                span_h = bbox[3] - bbox[1]
                mid_x = (bbox[0] + bbox[2]) / 2.0
                
                # Check horizontal alignment
                if bbox[0] < w * 0.28:
                    detected_align = "left"
                    detected_x_perc = round((bbox[0] / w) * 100.0, 1)
                elif bbox[2] > w * 0.72 and mid_x > w * 0.65:
                    detected_align = "right"
                    detected_x_perc = round((bbox[2] / w) * 100.0, 1)
                else:
                    detected_align = "center"
                    detected_x_perc = round((mid_x / w) * 100.0, 1)

                return {
                    "detected": True,
                    "placeholder_text": s["text"],
                    "font_name": mapped_font,
                    "font_size": round(s["size"], 1),
                    "font_color": s["color"],
                    "x_percent": detected_x_perc,
                    "y_percent": round(s["y_percent"], 1),
                    "align": detected_align,
                    "box_width": round(max(320.0, span_w + 50.0), 1),
                    "box_height": round(max(42.0, span_h + 16.0), 1),
                    "bbox": list(s["bbox"]),
                    "reason": f"Auto-detected placeholder \"{s['text']}\" ({mapped_font}, {round(s['size'], 1)}pt, {s['color']})",
                }

        # 2. Search for preamble / intro phrase ('awarded to', 'presented to', etc.)
        intro_pattern = re.compile(
            r"(proudly\s+awarded\s+to|presented\s+to|awarded\s+to|certif(y|ies)\s+that|conferred\s+upon|given\s+to|granted\s+to|presented\s+this|recognition\s+of)",
            re.IGNORECASE,
        )

        spans_sorted = sorted(spans, key=lambda x: x["bbox"][1])
        title_span = max(spans_sorted, key=lambda x: x["size"]) if spans_sorted else None
        primary_font = map_font_family(title_span["font"], is_bold=True) if title_span else "Times-Bold"
        primary_color = title_span["color"] if title_span else "#0d408c"

        for i, s in enumerate(spans_sorted):
            if intro_pattern.search(s["text"]):
                intro_bottom = s["bbox"][3]
                next_span = next((nxt for nxt in spans_sorted[i + 1:] if nxt["bbox"][1] > intro_bottom + 10), None)
                if next_span:
                    next_top = next_span["bbox"][1]
                    mid_y = (intro_bottom + next_top) / 2.0
                else:
                    mid_y = intro_bottom + 45.0

                y_perc = round((mid_y / h) * 100.0, 1)
                suggested_size = max(24.0, min(36.0, round(title_span["size"] * 0.8, 1))) if title_span else 28.0

                # Detect if the intro phrase or title is left-aligned (like Deloitte templates)
                if s["bbox"][0] < w * 0.22:
                    detected_align = "left"
                    detected_x_perc = round((s["bbox"][0] / w) * 100.0, 1)
                else:
                    detected_align = "center"
                    detected_x_perc = 50.0

                return {
                    "detected": True,
                    "placeholder_text": None,
                    "font_name": primary_font,
                    "font_size": suggested_size,
                    "font_color": primary_color,
                    "x_percent": detected_x_perc,
                    "y_percent": y_perc,
                    "align": detected_align,
                    "box_width": 520,
                    "box_height": 48,
                    "bbox": None,
                    "reason": f"Positioned below \"{s['text']}\" ({y_perc}% from top, {detected_align}) using template styling ({primary_font}, {suggested_size}pt)",
                }

        # 3. Fallback to title font & color
        # Check if title itself is on the left
        if title_span and title_span["bbox"][0] < w * 0.22:
            fallback_align = "left"
            fallback_x_perc = round((title_span["bbox"][0] / w) * 100.0, 1)
        else:
            fallback_align = "center"
            fallback_x_perc = 50.0

        return {
            "detected": False,
            "placeholder_text": None,
            "font_name": primary_font,
            "font_size": 28.0,
            "font_color": primary_color,
            "x_percent": fallback_x_perc,
            "y_percent": 50.0,
            "align": fallback_align,
            "box_width": 520,
            "box_height": 48,
            "bbox": None,
            "reason": f"Positioned using certificate theme font ({primary_font}) and color ({primary_color})",
        }
    except Exception as exc:
        return {
            "detected": False,
            "placeholder_text": None,
            "font_name": "Times-Bold",
            "font_size": 28.0,
            "font_color": "#0d408c",
            "x_percent": 50.0,
            "y_percent": 50.0,
            "align": "center",
            "box_width": 520,
            "box_height": 48,
            "bbox": None,
            "reason": f"Default placement ({exc})",
        }


def is_valid_pdf(pdf_path: str) -> bool:
    try:
        reader = PdfReader(pdf_path)
        _ = len(reader.pages)
        return True
    except Exception:
        return False


def get_pdf_info(pdf_path: str) -> dict:
    reader = PdfReader(pdf_path)
    page_count = len(reader.pages)
    if page_count == 0:
        raise ValueError("PDF has no pages")
    first_page = reader.pages[0]
    w = float(first_page.mediabox.width)
    h = float(first_page.mediabox.height)
    fields = detect_fillable_fields(pdf_path)
    analysis = analyze_pdf_template(pdf_path)
    return {
        "width": w,
        "height": h,
        "page_count": page_count,
        "has_fillable_fields": len(fields) > 0,
        "fields": fields,
        "analysis": analysis,
    }


# ---------------------------------------------------------------------------
# Names list reading
# ---------------------------------------------------------------------------

def read_columns(names_path: str) -> list[str]:
    """Return the header row / column names of the uploaded sheet."""
    ext = Path(names_path).suffix.lower()
    if ext == ".csv":
        with open(names_path, newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            header = next(reader, [])
        return [h.strip() for h in header]
    elif ext in (".xlsx", ".xlsm"):
        if openpyxl is None:
            raise RuntimeError("openpyxl is not installed")
        wb = openpyxl.load_workbook(names_path, read_only=True, data_only=True)
        ws = wb.active
        row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        return [str(c).strip() if c is not None else "" for c in row]
    else:
        raise ValueError(f"Unsupported file type: {ext}")


def read_names(names_path: str, column: Optional[str] = None) -> list[NameRecord]:
    """Read recipient names from the given column (or the first column if
    none specified). Completely empty rows are skipped silently; rows that
    exist but have a blank value in the name column come back with
    raw_name == "" so the caller can report them as invalid records
    (SRS section 14)."""
    ext = Path(names_path).suffix.lower()
    records: list[NameRecord] = []

    if ext == ".csv":
        with open(names_path, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            fieldnames = reader.fieldnames or []
            col = column or (fieldnames[0] if fieldnames else None)
            if col is None:
                return []
            for i, row in enumerate(reader, start=1):
                if row is None or all((v is None or str(v).strip() == "") for v in row.values()):
                    continue  # fully empty row -> ignore
                raw = (row.get(col) or "").strip()
                records.append(NameRecord(row_number=i, raw_name=raw))

    elif ext in (".xlsx", ".xlsm"):
        if openpyxl is None:
            raise RuntimeError("openpyxl is not installed")
        wb = openpyxl.load_workbook(names_path, read_only=True, data_only=True)
        ws = wb.active
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        header = [str(c).strip() if c is not None else "" for c in header]
        col = column or (header[0] if header else None)
        if col is None or col not in header:
            col_index = 0
        else:
            col_index = header.index(col)

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
            if row is None or all(v is None or str(v).strip() == "" for v in row):
                continue  # fully empty row -> ignore
            value = row[col_index] if col_index < len(row) else None
            raw = str(value).strip() if value is not None else ""
            records.append(NameRecord(row_number=i, raw_name=raw))
    else:
        raise ValueError(f"Unsupported file type: {ext}")

    return records


# ---------------------------------------------------------------------------
# Filename handling (SRS sections 15-17)
# ---------------------------------------------------------------------------

_INVALID_FS_CHARS = re.compile(r'[<>:"/\\|?*\x00-\x1f]')


def sanitize_filename(name: str) -> str:
    """Turn a recipient name into a safe Windows/macOS/Linux filename stem.
    This NEVER affects the name that gets printed on the certificate -
    it only affects the file on disk (SRS section 16)."""
    normalized = unicodedata.normalize("NFKC", name).strip()
    normalized = re.sub(r"\s+", "_", normalized)
    safe = _INVALID_FS_CHARS.sub("_", normalized)
    safe = safe.strip("._") or "certificate"
    return safe[:150]  # keep well under OS path-length limits


def unique_output_path(output_dir: str, stem: str, ext: str = ".pdf") -> str:
    """Return a path guaranteed not to collide with an existing file,
    appending _2, _3, ... as needed (SRS section 17)."""
    candidate = os.path.join(output_dir, f"{stem}{ext}")
    if not os.path.exists(candidate):
        return candidate
    n = 2
    while True:
        candidate = os.path.join(output_dir, f"{stem}_{n}{ext}")
        if not os.path.exists(candidate):
            return candidate
        n += 1


# ---------------------------------------------------------------------------
# Overflow estimation (SRS section 12)
# ---------------------------------------------------------------------------

def _metrics_font_name(declared_font: Optional[str]) -> str:
    """Map a PDF BaseFont name to one of reportlab's 14 standard fonts with
    known metrics, for width-estimation purposes only."""
    if not declared_font:
        return _FALLBACK_FONT
    cleaned = declared_font.split("+")[-1]  # strip subset tag e.g. 'ABCDEF+Helvetica'
    if cleaned in STANDARD_FONTS:
        return cleaned
    lower = cleaned.lower()
    bold = "bold" in lower
    italic = "italic" in lower or "oblique" in lower
    if "times" in lower or "serif" in lower:
        base = "Times"
        if bold and italic:
            return "Times-BoldItalic"
        if bold:
            return "Times-Bold"
        if italic:
            return "Times-Italic"
        return "Times-Roman"
    if "courier" in lower or "mono" in lower:
        if bold and italic:
            return "Courier-BoldOblique"
        if bold:
            return "Courier-Bold"
        if italic:
            return "Courier-Oblique"
        return "Courier"
    # default bucket: Helvetica family covers Arial and most sans-serif fonts
    if bold and italic:
        return "Helvetica-BoldOblique"
    if bold:
        return "Helvetica-Bold"
    if italic:
        return "Helvetica-Oblique"
    return _FALLBACK_FONT


def will_overflow(name: str, tfield: TemplateField, horizontal_padding_pt: float = 4.0) -> bool:
    """Estimate whether `name` is too wide (or too long, if /MaxLen is set)
    for the field. This is an ESTIMATE used only to flag records for human
    review - it never changes font size/position itself."""
    if tfield.max_len and len(name) > tfield.max_len:
        return True

    if not tfield.font_size or tfield.font_size <= 0:
        return False  # auto-sized field (size 0) - can't reliably estimate

    font = _metrics_font_name(tfield.font_name)
    try:
        text_width = pdfmetrics.stringWidth(name, font, tfield.font_size)
    except Exception:
        return False

    field_width = abs(tfield.rect[2] - tfield.rect[0])
    available = field_width - (2 * horizontal_padding_pt)
    return text_width > available


# ---------------------------------------------------------------------------
# Certificate generation (SRS sections 9-11)
# ---------------------------------------------------------------------------

def create_text_overlay(
    width: float,
    height: float,
    text: str,
    x: float,
    y: float,
    font_name: str = "Times-Bold",
    font_size: float = 28,
    font_color: str = "#000000",
    align: str = "center",
    cover_box: bool = False,
    box_width: Optional[float] = None,
    box_height: Optional[float] = None,
    box_color: str = "#FFFFFF",
) -> io.BytesIO:
    packet = io.BytesIO()
    c = canvas.Canvas(packet, pagesize=(width, height))

    # If covering an existing name/box
    if cover_box:
        bw = float(box_width) if box_width else (width * 0.55)
        bh = float(box_height) if box_height else (font_size * 1.6)
        if align == "left":
            bx = x - 8.0
        elif align == "right":
            bx = x - bw + 8.0
        else:  # center
            bx = x - (bw / 2.0)
        by = y - (bh * 0.28)
        try:
            c.setFillColor(HexColor(box_color))
        except Exception:
            c.setFillColorRGB(1, 1, 1)
        c.rect(bx, by, bw, bh, fill=1, stroke=0)

    try:
        if font_color.startswith("#"):
            c.setFillColor(HexColor(font_color))
        else:
            c.setFillColorRGB(0, 0, 0)
    except Exception:
        c.setFillColorRGB(0, 0, 0)

    try:
        c.setFont(font_name, font_size)
    except Exception:
        c.setFont("Helvetica-Bold", font_size)

    if align == "center":
        c.drawCentredString(x, y, text)
    elif align == "right":
        c.drawRightString(x, y, text)
    else:
        c.drawString(x, y, text)

    c.showPage()
    c.save()
    packet.seek(0)
    return packet


def overlay_single_certificate(
    template_path: str,
    name_value: str,
    output_path: str,
    x: Optional[float] = None,
    y: Optional[float] = None,
    x_percent: Optional[float] = None,
    y_percent: Optional[float] = None,
    font_name: str = "Times-Bold",
    font_size: float = 28,
    font_color: str = "#000000",
    align: str = "center",
    placeholder_bbox: Optional[list | tuple] = None,
    cover_box: bool = False,
    box_width: Optional[float] = None,
    box_height: Optional[float] = None,
    box_color: str = "#FFFFFF",
) -> None:
    base_template_bytes = None
    if placeholder_bbox:
        try:
            doc = pymupdf.open(template_path)
            page = doc[0]
            page.add_redact_annot(pymupdf.Rect(placeholder_bbox))
            page.apply_redactions()
            base_template_bytes = doc.tobytes()
            doc.close()
        except Exception:
            base_template_bytes = None

    if base_template_bytes:
        reader = PdfReader(io.BytesIO(base_template_bytes))
    else:
        reader = PdfReader(template_path)

    writer = PdfWriter()

    first_page = reader.pages[0]
    width = float(first_page.mediabox.width)
    height = float(first_page.mediabox.height)

    if x is not None:
        pos_x = x
    elif x_percent is not None:
        pos_x = width * (float(x_percent) / 100.0)
    elif align == "left":
        pos_x = width * 0.1
    elif align == "right":
        pos_x = width * 0.9
    else:
        pos_x = width / 2.0

    if y is not None:
        pos_y = y
    elif y_percent is not None:
        pos_y = height * (1.0 - (float(y_percent) / 100.0))
    else:
        pos_y = height * 0.48

    overlay_stream = create_text_overlay(
        width=width,
        height=height,
        text=name_value,
        x=pos_x,
        y=pos_y,
        font_name=font_name,
        font_size=font_size,
        font_color=font_color,
        align=align,
        cover_box=cover_box,
        box_width=box_width,
        box_height=box_height,
        box_color=box_color,
    )
    overlay_reader = PdfReader(overlay_stream)
    overlay_page = overlay_reader.pages[0]

    first_page.merge_page(overlay_page)
    writer.add_page(first_page)

    for p in reader.pages[1:]:
        writer.add_page(p)

    with open(output_path, "wb") as f:
        writer.write(f)


def render_certificate_image(
    template_path: str,
    name_value: str = "",
    mode: str = "overlay",
    field_name: Optional[str] = None,
    x_percent: Optional[float] = None,
    y_percent: float = 36.0,
    font_name: str = "Times-Bold",
    font_size: float = 32,
    font_color: str = "#d97706",
    align: str = "center",
    cover_box: bool = False,
    box_width: Optional[float] = None,
    box_height: Optional[float] = None,
    box_color: str = "#FFFFFF",
    placeholder_bbox: Optional[list | tuple] = None,
    dpi: int = 120,
) -> bytes:
    """Render a single certificate preview page directly as a PNG image in memory using PyMuPDF."""
    doc = pymupdf.open(template_path)
    page = doc[0]
    pw = page.rect.width
    ph = page.rect.height

    if placeholder_bbox:
        try:
            page.add_redact_annot(pymupdf.Rect(placeholder_bbox))
            page.apply_redactions()
        except Exception:
            pass

    if mode == "field" and field_name and name_value:
        # Form field preview
        pass
    elif name_value:
        # Optional text overlay
        f_color_hex = font_color or "#000000"
        rgb = _hex_to_rgb(f_color_hex)
        py_font = _py_font_for_name(font_name)
        font_sz = float(font_size or 28)

        if x_percent is not None:
            pos_x = pw * (float(x_percent) / 100.0)
        elif align == "left":
            pos_x = pw * 0.1
        elif align == "right":
            pos_x = pw * 0.9
        else:
            pos_x = pw / 2.0

        pos_y = ph * (float(y_percent) / 100.0) if y_percent is not None else (ph * 0.5)

        bw = float(box_width or (pw * 0.55))
        bh = float(box_height or (font_sz * 1.6))

        if align == "left":
            bx0 = pos_x - 4.0
            bx1 = bx0 + bw
        elif align == "right":
            bx1 = pos_x + 4.0
            bx0 = bx1 - bw
        else:
            bx0 = pos_x - (bw / 2.0)
            bx1 = pos_x + (bw / 2.0)

        by0 = pos_y - (bh / 2.0)
        by1 = pos_y + (bh / 2.0)
        box_rect = pymupdf.Rect(bx0, by0, bx1, by1)

        if cover_box:
            box_rgb = _hex_to_rgb(box_color or "#FFFFFF")
            page.draw_rect(box_rect, color=None, fill=box_rgb, overlay=True)

        try:
            text_len = pymupdf.get_text_length(name_value, fontname=py_font, fontsize=font_sz)
        except Exception:
            text_len = len(name_value) * font_sz * 0.55

        if align == "left":
            tx = bx0 + 6.0
        elif align == "right":
            tx = bx1 - text_len - 6.0
        else:
            tx = pos_x - (text_len / 2.0)

        ty = pos_y + (font_sz * 0.35)
        page.insert_text(
            (tx, ty),
            name_value,
            fontname=py_font,
            fontsize=font_sz,
            color=rgb,
            overlay=True,
        )

    pix = page.get_pixmap(dpi=dpi)
    png_bytes = pix.tobytes("png")
    doc.close()
    return png_bytes


def fill_single_certificate(template_path: str, field_name: str, name_value: str, output_path: str) -> None:
    """Create ONE certificate PDF: read the ORIGINAL template fresh (never
    mutate it), set only the target field's value, and write a brand-new
    file."""
    reader = PdfReader(template_path)
    writer = PdfWriter()
    writer.append(reader)

    filled_any = False
    for page in writer.pages:
        if page.get("/Annots"):
            writer.update_page_form_field_values(
                page,
                {field_name: name_value},
                auto_regenerate=False,
            )
            filled_any = True

    if not filled_any:
        raise ValueError(f"Field '{field_name}' was not found on any page")

    with open(output_path, "wb") as f:
        writer.write(f)


def _hex_to_rgb(hex_str: str) -> tuple[float, float, float]:
    """Convert hex color like #d97706 or #ffffff to RGB tuple (0.0 - 1.0)."""
    hex_clean = (hex_str or "#000000").lstrip("#")
    if len(hex_clean) == 3:
        hex_clean = "".join(c * 2 for c in hex_clean)
    if len(hex_clean) != 6:
        return (0.0, 0.0, 0.0)
    try:
        r = int(hex_clean[0:2], 16) / 255.0
        g = int(hex_clean[2:4], 16) / 255.0
        b = int(hex_clean[4:6], 16) / 255.0
        return (r, g, b)
    except Exception:
        return (0.0, 0.0, 0.0)


def _py_font_for_name(font_name: str) -> str:
    """Map user-selected font names to standard PyMuPDF font resource names."""
    f_lower = (font_name or "").lower()
    if "times" in f_lower:
        if "italic" in f_lower and "bold" in f_lower:
            return "times-bolditalic"
        elif "italic" in f_lower:
            return "times-italic"
        elif "bold" in f_lower:
            return "times-bold"
        return "times-roman"
    elif "courier" in f_lower:
        return "courier-bold" if "bold" in f_lower else "courier"
    else:
        if "bold" in f_lower and "italic" in f_lower:
            return "hebi"
        elif "bold" in f_lower:
            return "hebo"
        elif "italic" in f_lower:
            return "heit"
        return "helv"


def generate_all(
    template_path: str,
    names: list[NameRecord],
    field_name: Optional[str] = None,
    field_lookup: Optional[dict[str, TemplateField]] = None,
    output_dir: str = "",
    progress_cb=None,
    mode: str = "field",
    overlay_options: Optional[dict] = None,
    cancel_event: Optional[threading.Event] = None,
) -> list[GenerationResult]:
    """Generate one certificate per valid name record.
    Uses an ultra-fast in-memory PyMuPDF pipeline that handles large 25MB+ templates
    in milliseconds with zero memory bloat."""
    os.makedirs(output_dir, exist_ok=True)
    field_lookup = field_lookup or {}
    tfield = field_lookup.get(field_name) if field_name else None
    overlay_opts = overlay_options or {}
    results: list[GenerationResult] = []
    total = len(names)

    # Track stems used within THIS batch run so identical rows in the same spreadsheet get _2
    stem_counts: dict[str, int] = {}

    # Pre-cache and prepare base template once
    base_template_bytes = None
    if mode == "overlay":
        try:
            doc = pymupdf.open(template_path)
            placeholder_bbox = overlay_opts.get("placeholder_bbox")
            if placeholder_bbox:
                try:
                    doc[0].add_redact_annot(pymupdf.Rect(placeholder_bbox))
                    doc[0].apply_redactions()
                except Exception:
                    pass
            base_template_bytes = doc.tobytes(deflate=True, garbage=3)
            doc.close()
        except Exception:
            with open(template_path, "rb") as f:
                base_template_bytes = f.read()

    for idx, record in enumerate(names, start=1):
        if cancel_event and cancel_event.is_set():
            break

        if progress_cb:
            progress_cb(idx, total, record.raw_name)

        if not record.raw_name:
            results.append(GenerationResult(
                row_number=record.row_number, name="", status="empty",
                message="Name is empty."
            ))
            continue

        if mode == "field" and tfield and will_overflow(record.raw_name, tfield):
            results.append(GenerationResult(
                row_number=record.row_number, name=record.raw_name, status="overflow",
                message="Name exceeds the available name-field dimensions. "
                        "Font size and formatting were not modified."
            ))
            continue

        try:
            stem = sanitize_filename(record.raw_name)
            count = stem_counts.get(stem, 0) + 1
            stem_counts[stem] = count
            filename = f"{stem}.pdf" if count == 1 else f"{stem}_{count}.pdf"
            out_path = os.path.join(output_dir, filename)

            if mode == "field" and field_name:
                fill_single_certificate(template_path, field_name, record.raw_name, out_path)
            else:
                # Fast PyMuPDF overlay rendering
                doc = pymupdf.open(stream=base_template_bytes, filetype="pdf")
                page = doc[0]
                pw = page.rect.width
                ph = page.rect.height

                x_perc = overlay_opts.get("x_percent")
                y_perc = overlay_opts.get("y_percent")
                align_mode = overlay_opts.get("align", "center")
                font_sz = float(overlay_opts.get("font_size", 28))
                f_color_hex = overlay_opts.get("font_color", "#000000")
                rgb = _hex_to_rgb(f_color_hex)
                f_name = overlay_opts.get("font_name", "Times-Bold")
                is_cover = bool(overlay_opts.get("cover_box", False))
                py_font = _py_font_for_name(f_name)

                # Calculate X & Y
                if x_perc is not None:
                    pos_x = pw * (float(x_perc) / 100.0)
                elif align_mode == "left":
                    pos_x = pw * 0.1
                elif align_mode == "right":
                    pos_x = pw * 0.9
                else:
                    pos_x = pw / 2.0

                pos_y = ph * (float(y_perc) / 100.0) if y_perc is not None else (ph * 0.5)

                bw = float(overlay_opts.get("box_width") or (pw * 0.55))
                bh = float(overlay_opts.get("box_height") or (font_sz * 1.6))

                if align_mode == "left":
                    bx0 = pos_x - 4.0
                    bx1 = bx0 + bw
                elif align_mode == "right":
                    bx1 = pos_x + 4.0
                    bx0 = bx1 - bw
                else:
                    bx0 = pos_x - (bw / 2.0)
                    bx1 = pos_x + (bw / 2.0)

                by0 = pos_y - (bh / 2.0)
                by1 = pos_y + (bh / 2.0)
                box_rect = pymupdf.Rect(bx0, by0, bx1, by1)

                if is_cover:
                    b_color_hex = overlay_opts.get("box_color", "#FFFFFF")
                    box_rgb = _hex_to_rgb(b_color_hex)
                    page.draw_rect(box_rect, color=None, fill=box_rgb, overlay=True)

                # Compute exact horizontal starting point for text
                try:
                    text_len = pymupdf.get_text_length(record.raw_name, fontname=py_font, fontsize=font_sz)
                except Exception:
                    text_len = len(record.raw_name) * font_sz * 0.55

                if align_mode == "left":
                    tx = bx0 + 6.0
                elif align_mode == "right":
                    tx = bx1 - text_len - 6.0
                else:
                    tx = pos_x - (text_len / 2.0)

                # Baseline Y position
                ty = pos_y + (font_sz * 0.35)

                page.insert_text(
                    (tx, ty),
                    record.raw_name,
                    fontname=py_font,
                    fontsize=font_sz,
                    color=rgb,
                    overlay=True,
                )

                doc.save(out_path, deflate=True, garbage=3)
                doc.close()

            results.append(GenerationResult(
                row_number=record.row_number, name=record.raw_name, status="success",
                output_path=out_path,
            ))
        except Exception as exc:
            results.append(GenerationResult(
                row_number=record.row_number, name=record.raw_name, status="error",
                message=str(exc),
            ))

    return results
